import xlrd
from xlrd import open_workbook
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os

## Combines all of the information together using xlrd ## 
config_file_name = input("Insert the excel file name (without .xlsx): ")
config_sheet_name = input("Insert the sheet name: ")

print("Retreiving all of the items in the worksheet.....\n")

book = open_workbook(config_file_name+".xlsx")
sheet = book.sheet_by_name(config_sheet_name)

#first column: Index
index_array = sheet.col_values(0)
index_array.remove('index')

#second column: Service_Set_(new)
service_name_array = sheet.col_values(1)
service_name_array.remove('Service_Set_(new)')

#third column: TCP
tcp_array = sheet.col_values(2)
tcp_array.remove('TCP')
for i in range(len(tcp_array)):
    tcp_array[i] = tcp_array[i].split(',')

#fourth column: UDP
udp_array = sheet.col_values(3)
udp_array.remove('UDP')
for i in range(len(udp_array)):
    udp_array[i] = udp_array[i].split(',')


#fifth column: description
desc_array = sheet.col_values(4)
desc_array.remove('description')

size = len(index_array)

config_array = []

for i in range(size):
    config_array.append([])
    config_array[i].append("#")
    if(tcp_array[i][0]==""):
        protocol_type = "udp"
        protocol_array = udp_array[i]
    else:
        protocol_type = "tcp"
        protocol_array = tcp_array[i]
    config = "ip service-set "+service_name_array[i]+" type object"
    config_array[i].append(config)
    if(desc_array[i]!=""):
        desc_line = "    description "+ desc_array[i]
        config_array[i].append(desc_line)

    for j in range(len(protocol_array)):
        protocol_array[j]=str(protocol_array[j]).split("-")
        if(len(protocol_array[j])==2):
            service_code = "    service "+str(j)+" protocol "+protocol_type+" destination-port "+protocol_array[j][0]+" to "+protocol_array[j][1]
        else:
            service_code = "    service "+str(j)+" protocol "+protocol_type+" destination-port "+protocol_array[j][0]
        config_array[i].append(service_code)
    config_array[i].append("#")

## Reads the information and saves into a new file using openpyxl ##

dest_file = input("Input your destination file name (without .xlsx): ")
dest_file2 = "only_config.xlsx"
print("Saving all items in the excel file....")
print("You should have 2 files in the folder, namely:")
print("[ "+dest_file+".xlsx ] which adds a column to the original file, and")
print("[ "+dest_file2+" ] containing only a configuration column.\n")

workbook = load_workbook("config_file.xlsx")
workbook2 = Workbook()
worksheet = workbook.get_sheet_by_name(config_sheet_name)
worksheet2 = workbook2.active

for x in range(len(index_array)):
    write_string = ''
    for j in range(len(config_array[x])):
        write_string = write_string + config_array[x][j]+"\n"
    worksheet['F1']='Configuration'
    worksheet2['A1']='Configuration'
    worksheet['F'+str(x+2)]=write_string
    worksheet2['A'+str(x+2)]=write_string

workbook.save(dest_file+".xlsx")
workbook.close()

workbook2.save(dest_file2)
workbook.close()
print("The program has finished processing.")
os.system('pause')
